#!/usr/bin/env python3
"""
CoG Decel AE = Above Expected = 실제 CoG_Decel − 예측 CoG_Decel(구속 기반)
- KR cohort xlsx 로 회귀: CoG_Decel = a × ball_speed + b
- TestPlayer의 AE = 실제 CoG_Decel − 예측값
"""
import os, sys, json, statistics
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'sample_data/KR_pitching_processing.xlsx')
TEST_JSON = os.path.join(ROOT, 'sample_data/theia_TEST_9trial_single.json')

wb = load_workbook(XLSX, data_only=True)
ws = wb.active
header = list(next(ws.iter_rows(values_only=True)))
ix_velo = header.index('ball_speed')
ix_cog = header.index('CoG_Decel')

# KR cohort 데이터 수집
data = []
for row in list(ws.iter_rows(values_only=True))[1:]:
    bs = row[ix_velo]
    cd = row[ix_cog]
    if bs is None or cd is None: continue
    try:
        bs = float(bs); cd = float(cd)
    except:
        continue
    # outlier 제거 (xlsx ball_speed 단위는 km/h)
    if 110 <= bs <= 160 and 0.5 <= cd <= 3.0:
        data.append((bs, cd))

n = len(data)
print(f"KR cohort 유효 trial: {n} (xlsx 가공값)")
print(f"ball_speed 범위 (km/h): {min(b for b,_ in data):.1f} ~ {max(b for b,_ in data):.1f}")
print(f"CoG_Decel 범위 (m/s): {min(c for _,c in data):.2f} ~ {max(c for _,c in data):.2f}")

# 단순 선형 회귀 (statistics)
xs = [b for b,_ in data]
ys = [c for _,c in data]
mx = statistics.mean(xs); my = statistics.mean(ys)
sx2 = sum((x-mx)**2 for x in xs)
sxy = sum((x-mx)*(y-my) for x,y in zip(xs,ys))
a = sxy/sx2     # slope
b0 = my - a*mx  # intercept
print(f"\n회귀식: CoG_Decel = {a:.4f} × ball_speed_kmh + {b0:.4f}")
print(f"  cohort 평균 ball_speed: {mx:.1f} km/h, 평균 CoG_Decel: {my:.2f} m/s")

# 회귀 잔차 통계 → AE의 cohort 분포
residuals = [c - (a*b + b0) for b,c in data]
res_sd = statistics.stdev(residuals)
res_med = statistics.median(residuals)
print(f"  잔차 통계: median={res_med:.3f}, sd={res_sd:.3f} m/s")

# TestPlayer 적용 (140 km/h ≈ 87 mph, CoG_Decel = 1.77 m/s)
test_velocity_kmh = 141.1   # measured
test_cog_decel = 1.77

predicted = a * test_velocity_kmh + b0
ae = test_cog_decel - predicted
print(f"\n[TestPlayer 적용]")
print(f"  ball_speed   = {test_velocity_kmh:.1f} km/h")
print(f"  CoG_Decel    = {test_cog_decel:.2f} m/s (실제)")
print(f"  Predicted    = {predicted:.2f} m/s (회귀선)")
print(f"  CoG_Decel AE = {ae:+.3f} m/s (실제 − 예측)")

# JSON 갱신
with open(TEST_JSON, 'r', encoding='utf-8') as f:
    rec = json.load(f)
rec.setdefault('cog', {})['decel_ae'] = round(ae, 3)
rec['cog']['decel_ae_method'] = f'KR cohort regression: CoG_Decel = {a:.4f} × ball_speed_kmh + {b0:.4f}'
rec['cog']['decel_ae_predicted'] = round(predicted, 3)
with open(TEST_JSON, 'w', encoding='utf-8') as f:
    json.dump(rec, f, ensure_ascii=False, indent=2)
print(f"\n✓ {os.path.basename(TEST_JSON)} 갱신 (cog.decel_ae 추가)")
